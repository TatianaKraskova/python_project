import openpyxl

# Load workbook and sheet
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Initialize dictionaries
product_per_supplier = {}
total_value_per_supplier = {}
product_under_50 = {}

print("Max Rows:", product_list.max_row)
print("Max Columns:", product_list.max_column)

# Iterate through each product row
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value  # Column 4 = Supplier
    inventory = product_list.cell(product_row, 2).value      # Column 2 = Inventory
    price = product_list.cell(product_row, 3).value          # Column 3 = Price
    product_num = product_list.cell(product_row, 1).value    # Column 1 = Product Number
    inventory_price = product_list.cell(product_row, 5)

    # Skip rows with missing supplier name or invalid inventory/price
    if supplier_name is None or inventory is None or price is None:
        continue

    # Count products per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

    # Calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Identify products with inventory less than 50
    if inventory < 50:
        product_under_50[int(product_num)] = int(inventory)
    # Add value for the inventory_price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")

# Final Results
print("Product Count Per Supplier:", product_per_supplier)
print("Total Inventory Value Per Supplier:", total_value_per_supplier)
print("Products with Inventory Less than 50:", product_under_50)
print("Products with total price:", inventory_price.value)
